import sqlite3
import os

from PyQt6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QLabel,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QPushButton,
    QHBoxLayout,
    QMessageBox,
    QDialog,
    QFormLayout,
    QLineEdit,
    QComboBox,
    QFileDialog
)

from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QColor

try:
    from openpyxl import Workbook, load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


DB_PATH = os.path.join(
    os.path.dirname(__file__),
    "database/logistics.db"
)

class EditOrderDialog(QDialog):
    """编辑订单对话框"""

    def __init__(self, order_id, customer_name, x, y, demand, status):
        super().__init__()

        self.order_id = order_id

        self.setWindowTitle(f"✏️ 修改订单 - {order_id}")
        self.resize(400, 300)

        layout = QFormLayout()
        self.setLayout(layout)

        # ================= 输入框 =================
        self.customer_edit = QLineEdit(customer_name)

        self.x_edit = QLineEdit(str(x))

        self.y_edit = QLineEdit(str(y))

        self.demand_edit = QLineEdit(str(demand))

        self.status_combo = QComboBox()
        self.status_combo.addItems(["pending", "completed"])
        self.status_combo.setCurrentText(status)

        layout.addRow("客户姓名：", self.customer_edit)
        layout.addRow("X坐标：", self.x_edit)
        layout.addRow("Y坐标：", self.y_edit)
        layout.addRow("需求量：", self.demand_edit)
        layout.addRow("状态：", self.status_combo)

        # ================= 按钮 =================
        btn_layout = QHBoxLayout()

        self.save_btn = QPushButton("💾 保存修改")
        self.save_btn.clicked.connect(self.save_order)

        self.cancel_btn = QPushButton("❌ 取消")
        self.cancel_btn.clicked.connect(self.reject)

        btn_layout.addWidget(self.save_btn)
        btn_layout.addWidget(self.cancel_btn)

        layout.addRow(btn_layout)

        self.setStyleSheet("""
            QWidget{
                background-color:#0f172a;
                color:white;
                font-size:14px;
            }
            QLineEdit,QComboBox{
                background-color:#1e293b;
                border:1px solid #334155;
                padding:6px;
                border-radius:6px;
                color:white;
            }
            QPushButton{
                background-color:#2563eb;
                padding:8px;
                border-radius:8px;
                color:white;
            }
            QPushButton:hover{
                background-color:#3b82f6;
            }
        """)

    def save_order(self):
        customer = self.customer_edit.text().strip()
        x = self.x_edit.text().strip()
        y = self.y_edit.text().strip()
        demand = self.demand_edit.text().strip()
        status = self.status_combo.currentText()

        if not all([customer, x, y, demand]):
            QMessageBox.warning(
                self,
                "错误",
                "请填写完整信息"
            )
            return

        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()

            cursor.execute("""
                UPDATE orders
                SET customer_name=?,
                    x=?,
                    y=?,
                    demand=?,
                    status=?
                WHERE id=?
            """, (
                customer,
                float(x),
                float(y),
                float(demand),
                status,
                self.order_id
            ))

            conn.commit()
            conn.close()

            QMessageBox.information(
                self,
                "成功",
                "订单修改成功"
            )

            self.accept()

        except Exception as e:
            QMessageBox.warning(
                self,
                "错误",
                str(e)
            )


class AddOrderDialog(QDialog):

    def __init__(self):
        super().__init__()

        self.setWindowTitle("➕ 新增订单")

        self.resize(400, 300)

        layout = QFormLayout()

        self.setLayout(layout)

        # ================= 输入框 =================
        self.customer_edit = QLineEdit()

        self.x_edit = QLineEdit()

        self.y_edit = QLineEdit()

        self.demand_edit = QLineEdit()

        self.status_combo = QComboBox()

        self.status_combo.addItems([
            "pending",
            "completed"
        ])

        layout.addRow("客户姓名：", self.customer_edit)

        layout.addRow("X坐标：", self.x_edit)

        layout.addRow("Y坐标：", self.y_edit)

        layout.addRow("需求量：", self.demand_edit)

        layout.addRow("状态：", self.status_combo)

        # ================= 按钮 =================
        self.save_btn = QPushButton("💾 保存订单")

        self.save_btn.clicked.connect(self.save_order)

        layout.addRow(self.save_btn)

        self.setStyleSheet("""
            QWidget{
                background-color:#0f172a;
                color:white;
                font-size:14px;
            }

            QLineEdit,QComboBox{
                background-color:#1e293b;
                border:1px solid #334155;
                padding:6px;
                border-radius:6px;
                color:white;
            }

            QPushButton{
                background-color:#2563eb;
                padding:8px;
                border-radius:8px;
                color:white;
            }

            QPushButton:hover{
                background-color:#3b82f6;
            }
        """)

    def save_order(self):
        customer = self.customer_edit.text().strip()
        x = self.x_edit.text().strip()
        y = self.y_edit.text().strip()
        demand = self.demand_edit.text().strip()
        status = self.status_combo.currentText()

        if not all([customer, x, y, demand]):
            QMessageBox.warning(self, "错误", "请填写完整信息")
            return

        try:
            from datetime import datetime
            create_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()

            cursor.execute("""
                INSERT INTO orders
                (customer_name, x, y, demand, status, create_time)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (customer, float(x), float(y), float(demand), status, create_time))

            conn.commit()
            conn.close()

            QMessageBox.information(self, "成功", "订单添加成功")
            self.accept()

        except Exception as e:

            QMessageBox.warning(
                self,
                "错误",
                str(e)
            )
class OrderManagerWindow(QWidget):

    def __init__(self):
        super().__init__()

        self.setWindowTitle("📦 订单管理系统")
        self.resize(1000, 600)

        self.init_ui()

        self.load_orders()

    def init_ui(self):

        main_layout = QVBoxLayout()

        self.setLayout(main_layout)

        # ================= 标题 =================
        title = QLabel("📦 订单管理中心")
        title.setFont(QFont("Microsoft YaHei", 16, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("""
            color: white;
            padding: 10px;
            background-color: #1e293b;
            border-radius: 10px;
        """)
        main_layout.addWidget(title)

        # ================= 搜索区域 =================
        search_layout = QHBoxLayout()
        
        search_label = QLabel("🔍 搜索：")
        search_label.setStyleSheet("color: white; font-size: 14px;")
        search_layout.addWidget(search_label)
        
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("输入客户姓名或订单号...")
        self.search_edit.setStyleSheet("""
            QLineEdit{
                background-color:#1e293b;
                border:1px solid #334155;
                padding:8px;
                border-radius:6px;
                color:white;
                font-size:14px;
            }
        """)
        self.search_edit.returnPressed.connect(self.search_orders)
        search_layout.addWidget(self.search_edit, 1)
        
        self.search_btn = QPushButton("🔍 搜索")
        self.search_btn.clicked.connect(self.search_orders)
        self.search_btn.setStyleSheet("""
            QPushButton{
                background-color:#3b82f6;
                color:white;
                padding:8px 16px;
                border-radius:8px;
                font-size:14px;
            }
            QPushButton:hover{
                background-color:#60a5fa;
            }
        """)
        search_layout.addWidget(self.search_btn)
        
        self.clear_search_btn = QPushButton("❌ 清空")
        self.clear_search_btn.clicked.connect(self.clear_search)
        self.clear_search_btn.setStyleSheet("""
            QPushButton{
                background-color:#64748b;
                color:white;
                padding:8px 16px;
                border-radius:8px;
                font-size:14px;
            }
            QPushButton:hover{
                background-color:#94a3b8;
            }
        """)
        search_layout.addWidget(self.clear_search_btn)
        
        main_layout.addLayout(search_layout)
        
        # ================= Excel导入导出区域 =================
        excel_layout = QHBoxLayout()
        
        self.export_btn = QPushButton("📤 导出Excel")
        self.export_btn.clicked.connect(self.export_to_excel)
        self.export_btn.setStyleSheet("""
            QPushButton{
                background-color:#10b981;
                color:white;
                padding:8px 16px;
                border-radius:8px;
                font-size:14px;
            }
            QPushButton:hover{
                background-color:#34d399;
            }
        """)
        excel_layout.addWidget(self.export_btn)
        
        self.import_btn = QPushButton("📥 导入Excel")
        self.import_btn.clicked.connect(self.import_from_excel)
        self.import_btn.setStyleSheet("""
            QPushButton{
                background-color:#8b5cf6;
                color:white;
                padding:8px 16px;
                border-radius:8px;
                font-size:14px;
            }
            QPushButton:hover{
                background-color:#a78bfa;
            }
        """)
        excel_layout.addWidget(self.import_btn)
        
        excel_layout.addStretch()
        
        main_layout.addLayout(excel_layout)

        # ================= 按钮区域 =================
        btn_layout = QHBoxLayout()

        self.refresh_btn = QPushButton("🔄 刷新订单")
        # ================= 新增订单按钮 =================
        self.add_btn = QPushButton("➕ 新增订单")

        self.add_btn.clicked.connect(self.add_order)

        self.add_btn.setStyleSheet("""
            QPushButton{
                background-color:#16a34a;
                color:white;
                padding:8px;
                border-radius:8px;
                font-size:14px;
            }

            QPushButton:hover{
                background-color:#22c55e;
            }
        """)

        btn_layout.addWidget(self.add_btn)

        # ================= 删除订单按钮 =================
        self.delete_btn = QPushButton("❌ 删除订单")

        self.delete_btn.clicked.connect(self.delete_order)

        self.delete_btn.setStyleSheet("""
            QPushButton{
                background-color:#dc2626;
                color:white;
                padding:8px;
                border-radius:8px;
                font-size:14px;
            }

            QPushButton:hover{
                background-color:#ef4444;
            }
        """)

        btn_layout.addWidget(self.delete_btn)
        
        # ================= 修改订单按钮 =================
        self.edit_btn = QPushButton("✏️ 修改订单")
        self.edit_btn.clicked.connect(self.edit_order)
        self.edit_btn.setStyleSheet("""
            QPushButton{
                background-color:#f59e0b;
                color:white;
                padding:8px;
                border-radius:8px;
                font-size:14px;
            }
            QPushButton:hover{
                background-color:#fbbf24;
            }
        """)
        btn_layout.addWidget(self.edit_btn)

        self.refresh_btn.clicked.connect(self.load_orders)

        self.refresh_btn.setStyleSheet("""
            QPushButton{
                background-color:#2563eb;
                color:white;
                padding:8px;
                border-radius:8px;
                font-size:14px;
            }

            QPushButton:hover{
                background-color:#3b82f6;
            }
        """)

        btn_layout.addWidget(self.refresh_btn)
        
        # ================= 批量删除按钮 =================
        self.batch_delete_btn = QPushButton("🗑️ 批量删除")
        self.batch_delete_btn.clicked.connect(self.batch_delete_orders)
        self.batch_delete_btn.setStyleSheet("""
            QPushButton{
                background-color:#7c3aed;
                color:white;
                padding:8px;
                border-radius:8px;
                font-size:14px;
            }
            QPushButton:hover{
                background-color:#8b5cf6;
            }
        """)
        btn_layout.addWidget(self.batch_delete_btn)

        btn_layout.addStretch()

        main_layout.addLayout(btn_layout)

        # ================= 表格 =================
        self.table = QTableWidget()

        self.table.setColumnCount(8)

        self.table.setHorizontalHeaderLabels([
            "订单号",
            "客户",
            "下单时间",
            "发车时间",
            "完成时间",
            "耗时",
            "效率",
            "状态"
        ])

        # 表头自适应
        self.table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch
        )

        # 不允许编辑
        self.table.setEditTriggers(
            QTableWidget.EditTrigger.NoEditTriggers
        )

        # 整行选中
        self.table.setSelectionBehavior(
            QTableWidget.SelectionBehavior.SelectRows
        )
        
        # 双击编辑
        self.table.doubleClicked.connect(self.on_table_double_clicked)

        self.table.setStyleSheet("""
            QTableWidget{
                background-color:#0f172a;
                color:white;
                gridline-color:#334155;
                font-size:13px;
            }

            QHeaderView::section{
                background-color:#1e293b;
                color:white;
                padding:8px;
                border:none;
                font-weight:bold;
            }

            QTableWidget::item:selected{
                background-color:#2563eb;
            }
        """)

        main_layout.addWidget(self.table)

        self.setStyleSheet("""
            QWidget{
                background-color:#020617;
            }
        """)
    def add_order(self):

        dialog = AddOrderDialog()

        if dialog.exec():

            self.load_orders()

    def search_orders(self):
        """搜索订单"""
        keyword = self.search_edit.text().strip()
        
        if not keyword:
            self.load_orders()
            return
        
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            
            # 使用 LIKE 进行模糊查询
            cursor.execute("""
                SELECT
                    id,
                    customer_name,
                    x,
                    y,
                    demand,
                    status
                FROM orders
                WHERE customer_name LIKE ? OR CAST(id AS TEXT) LIKE ?
            """, (f'%{keyword}%', f'%{keyword}%'))
            
            orders = cursor.fetchall()
            conn.close()
            
            self.table.setRowCount(len(orders))
            
            for row_idx, row_data in enumerate(orders):
                for col_idx, value in enumerate(row_data):
                    item = QTableWidgetItem(str(value))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.table.setItem(row_idx, col_idx, item)
            
            self.status_msg = QLabel(f"🔍 搜索 '{keyword}' 找到 {len(orders)} 条结果")
            
        except Exception as e:
            QMessageBox.warning(
                self,
                "错误",
                f"搜索失败：\n{str(e)}"
            )

    def clear_search(self):
        """清空搜索"""
        self.search_edit.clear()
        self.load_orders()

    def export_to_excel(self):
        """导出订单到 Excel"""
        if not EXCEL_AVAILABLE:
            QMessageBox.warning(
                self,
                "错误",
                "请先安装 openpyxl 库：\npip install openpyxl"
            )
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出订单",
            "orders.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, customer_name, x, y, demand, status
                FROM orders
            """)
            orders = cursor.fetchall()
            conn.close()
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "订单列表"
            
            # 写入表头
            headers = ["订单号", "客户姓名", "X坐标", "Y坐标", "需求量", "状态"]
            ws.append(headers)
            
            # 写入数据
            for order in orders:
                ws.append(order)
            
            # 调整列宽
            column_widths = [10, 15, 10, 10, 10, 10]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + i)].width = width
            
            wb.save(file_path)
            
            QMessageBox.information(
                self,
                "成功",
                f"成功导出 {len(orders)} 条订单到：\n{file_path}"
            )
            
        except Exception as e:
            QMessageBox.warning(
                self,
                "错误",
                f"导出失败：\n{str(e)}"
            )

    def import_from_excel(self):
        """从 Excel 导入订单 - 同步到调度系统"""
        if not EXCEL_AVAILABLE:
            QMessageBox.warning(
                self,
                "错误",
                "请先安装 openpyxl 库：\npip install openpyxl"
            )
            return
        
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "导入订单",
            "",
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 跳过表头，从第二行开始读取
            imported_count = 0
            error_count = 0
            
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    # 检查数据完整性
                    if len(row) < 5:
                        continue
                    
                    customer_name = str(row[1]) if row[1] else ""
                    x = float(row[2]) if row[2] else 0
                    y = float(row[3]) if row[3] else 0
                    demand = float(row[4]) if row[4] else 0
                    status = str(row[5]) if len(row) > 5 and row[5] else "pending"
                    
                    if not customer_name or demand <= 0:
                        continue
                    
                    # 插入到 orders 表
                    from datetime import datetime
                    create_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    cursor.execute("""
                        INSERT INTO orders (customer_name, x, y, demand, status, create_time)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (customer_name, x, y, demand, status, create_time))
                    
                    # 同步插入到 customers 表（调度系统使用）
                    # 生成唯一ID
                    nid = f"C{customer_name}"
                    counter = 1
                    cursor.execute("SELECT id FROM customers WHERE id=?", (nid,))
                    while cursor.fetchone():
                        nid = f"C{customer_name}_{counter}"
                        counter += 1
                        cursor.execute("SELECT id FROM customers WHERE id=?", (nid,))
                    
                    cursor.execute("""
                        INSERT OR REPLACE INTO customers (id, x, y, q)
                        VALUES (?, ?, ?, ?)
                    """, (nid, x, y, demand))
                    
                    imported_count += 1
                    
                except Exception as e:
                    error_count += 1
                    print(f"导入行失败: {row}, 错误: {e}")
            
            conn.commit()
            conn.close()
            
            self.load_orders()
            
            # 发送信号通知主窗口刷新数据
            self.on_import_completed()
            
            msg = f"成功导入 {imported_count} 条订单"
            if error_count > 0:
                msg += f"\n{error_count} 条记录导入失败"
            
            QMessageBox.information(
                self,
                "完成",
                msg + "\n调度系统数据已同步更新"
            )
            
        except Exception as e:
            QMessageBox.warning(
                self,
                "错误",
                f"导入失败：\n{str(e)}"
            )
    
    def on_import_completed(self):
        """导入完成后的回调 - 子类可重写此方法通知主窗口"""
        pass

    def on_table_double_clicked(self, index):
        """双击表格打开编辑对话框"""
        self.edit_order()

    def edit_order(self):
        """修改订单"""
        current_row = self.table.currentRow()

        if current_row == -1:
            QMessageBox.warning(
                self,
                "提示",
                "请先选择要修改的订单"
            )
            return

        # 获取当前行的数据
        order_id = self.table.item(current_row, 0).text()
        customer_name = self.table.item(current_row, 1).text()
        x = self.table.item(current_row, 2).text()
        y = self.table.item(current_row, 3).text()
        demand = self.table.item(current_row, 4).text()
        status = self.table.item(current_row, 5).text()

        # 打开编辑对话框
        dialog = EditOrderDialog(order_id, customer_name, x, y, demand, status)
        if dialog.exec():
            self.load_orders()
    def delete_order(self):

        current_row = self.table.currentRow()

        if current_row == -1:

            QMessageBox.warning(
                self,
                "提示",
                "请先选择订单"
            )

            return

        order_id = self.table.item(current_row, 0).text()

        reply = QMessageBox.question(
            self,
            "确认删除",
            f"确定删除订单 {order_id} 吗？",
            QMessageBox.StandardButton.Yes |
            QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:

            try:

                conn = sqlite3.connect(DB_PATH)

                cursor = conn.cursor()

                cursor.execute("""
                    DELETE FROM orders
                    WHERE id=?
                """, (order_id,))

                conn.commit()

                conn.close()

                QMessageBox.information(
                    self,
                    "成功",
                    "订单删除成功"
                )

                self.load_orders()

            except Exception as e:

                QMessageBox.warning(
                    self,
                    "错误",
                    str(e)
                )
    def load_orders(self):
        """加载订单并显示时间数据"""
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()

            cursor.execute("""
                SELECT
                    id,
                    customer_name,
                    create_time,
                    delivery_time,
                    finish_time,
                    duration,
                    efficiency,
                    status,
                    demand
                FROM orders
            """)

            orders = cursor.fetchall()
            conn.close()

            self.table.setRowCount(len(orders))

            for row_idx, row_data in enumerate(orders):
                (order_id, customer_name, create_time, delivery_time, 
                 finish_time, duration, efficiency, status, demand) = row_data
                
                # 计算耗时和效率（如果有时间数据）
                if delivery_time and finish_time:
                    try:
                        from datetime import datetime
                        fmt = "%Y-%m-%d %H:%M:%S"
                        start = datetime.strptime(delivery_time, fmt)
                        end = datetime.strptime(finish_time, fmt)
                        minutes = int((end - start).total_seconds() / 60)
                        duration_str = f"{minutes} 分钟"
                        
                        # 效率 = 重量 / 时间
                        if minutes > 0 and demand:
                            eff = round(demand / minutes, 2)
                            efficiency_str = f"{eff} kg/min"
                        else:
                            efficiency_str = "-"
                    except:
                        duration_str = duration if duration else "-"
                        efficiency_str = efficiency if efficiency else "-"
                else:
                    duration_str = "-"
                    efficiency_str = "-"
                
                # 格式化时间显示
                create_time_str = create_time if create_time else "-"
                delivery_time_str = delivery_time if delivery_time else "-"
                finish_time_str = finish_time if finish_time else "-"
                
                # 准备数据
                data = [
                    str(order_id),
                    customer_name,
                    create_time_str,
                    delivery_time_str,
                    finish_time_str,
                    duration_str,
                    efficiency_str,
                    status
                ]
                
                # 插入到表格
                for col, text in enumerate(data):
                    item = QTableWidgetItem(str(text))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    
                    # 根据状态设置颜色
                    if col == 7:  # 状态列
                        if status == "已完成":
                            item.setBackground(QColor("#2ecc71"))
                            item.setForeground(QColor("#ffffff"))
                        elif status == "配送中":
                            item.setBackground(QColor("#3498db"))
                            item.setForeground(QColor("#ffffff"))
                        elif status == "超时":
                            item.setBackground(QColor("#e74c3c"))
                            item.setForeground(QColor("#ffffff"))
                    
                    self.table.setItem(row_idx, col, item)

        except Exception as e:
            QMessageBox.warning(
                self,
                "错误",
                f"加载订单失败：\n{str(e)}"
            )

    def batch_delete_orders(self):
        """批量删除选中的订单"""
        # 获取选中的行
        selected_rows = set()
        for item in self.table.selectedItems():
            selected_rows.add(item.row())
        
        if not selected_rows:
            QMessageBox.warning(
                self,
                "提示",
                "请先选择要删除的订单（可多选）"
            )
            return
        
        # 获取选中的订单ID
        order_ids = []
        for row in selected_rows:
            order_id = self.table.item(row, 0).text()
            order_ids.append(order_id)
        
        # 确认删除
        reply = QMessageBox.question(
            self,
            "确认批量删除",
            f"确定要删除选中的 {len(order_ids)} 个订单吗？\n\n订单号: {', '.join(order_ids[:5])}{'...' if len(order_ids) > 5 else ''}",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                conn = sqlite3.connect(DB_PATH)
                cursor = conn.cursor()
                
                # 批量删除
                deleted_count = 0
                for order_id in order_ids:
                    cursor.execute("""
                        DELETE FROM orders
                        WHERE id=?
                    """, (order_id,))
                    deleted_count += 1
                
                conn.commit()
                conn.close()
                
                QMessageBox.information(
                    self,
                    "成功",
                    f"成功删除 {deleted_count} 个订单"
                )
                
                # 刷新表格
                self.load_orders()
                
            except Exception as e:
                QMessageBox.warning(
                    self,
                    "错误",
                    f"批量删除失败：\n{str(e)}"
                )


class SystemLogWindow(QWidget):
    """系统日志窗口"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📋 系统日志")
        self.resize(800, 600)
        
        layout = QVBoxLayout()
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(10)
        self.setLayout(layout)
        
        # 标题
        title = QLabel("📋 系统操作日志")
        title.setFont(QFont("Microsoft YaHei", 16, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("""
            color: white;
            padding: 10px;
            background-color: #1e293b;
            border-radius: 10px;
        """)
        layout.addWidget(title)
        
        # 日志表格
        self.log_table = QTableWidget()
        self.log_table.setColumnCount(4)
        self.log_table.setHorizontalHeaderLabels(["时间", "用户", "操作类型", "详情"])
        self.log_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.log_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.log_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.log_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.log_table.setStyleSheet("""
            QTableWidget{
                background-color:#0f172a;
                color:white;
                gridline-color:#334155;
                font-size:13px;
            }
            QHeaderView::section{
                background-color:#1e293b;
                color:white;
                padding:8px;
                border:none;
                font-weight:bold;
            }
            QTableWidget::item:selected{
                background-color:#2563eb;
            }
        """)
        layout.addWidget(self.log_table)
        
        # 按钮区域
        btn_layout = QHBoxLayout()
        
        refresh_btn = QPushButton("🔄 刷新")
        refresh_btn.clicked.connect(self.load_logs)
        refresh_btn.setStyleSheet("""
            QPushButton{
                background-color:#3b82f6;
                color:white;
                padding:8px 16px;
                border-radius:8px;
                font-size:14px;
            }
            QPushButton:hover{
                background-color:#60a5fa;
            }
        """)
        btn_layout.addWidget(refresh_btn)
        
        clear_btn = QPushButton("🗑️ 清空日志")
        clear_btn.clicked.connect(self.clear_logs)
        clear_btn.setStyleSheet("""
            QPushButton{
                background-color:#dc2626;
                color:white;
                padding:8px 16px;
                border-radius:8px;
                font-size:14px;
            }
            QPushButton:hover{
                background-color:#ef4444;
            }
        """)
        btn_layout.addWidget(clear_btn)
        
        btn_layout.addStretch()
        
        close_btn = QPushButton("✕ 关闭")
        close_btn.clicked.connect(self.close)
        close_btn.setStyleSheet("""
            QPushButton{
                background-color:#64748b;
                color:white;
                padding:8px 16px;
                border-radius:8px;
                font-size:14px;
            }
            QPushButton:hover{
                background-color:#94a3b8;
            }
        """)
        btn_layout.addWidget(close_btn)
        
        layout.addLayout(btn_layout)
        
        # 设置窗口样式
        self.setStyleSheet("""
            QWidget{
                background-color:#020617;
            }
        """)
        
        # 加载日志
        self.load_logs()
    
    def load_logs(self):
        """从数据库加载日志"""
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            
            # 获取最近的100条日志
            cursor.execute("""
                SELECT timestamp, username, action
                FROM logs
                ORDER BY timestamp DESC
                LIMIT 100
            """)
            
            logs = cursor.fetchall()
            conn.close()
            
            self.log_table.setRowCount(len(logs))
            
            for row_idx, (timestamp, username, action) in enumerate(logs):
                # 时间
                time_item = QTableWidgetItem(str(timestamp))
                time_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.log_table.setItem(row_idx, 0, time_item)
                
                # 用户
                user_item = QTableWidgetItem(str(username))
                user_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.log_table.setItem(row_idx, 1, user_item)
                
                # 操作类型（从action中提取）
                action_type = self._get_action_type(action)
                type_item = QTableWidgetItem(action_type)
                type_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.log_table.setItem(row_idx, 2, type_item)
                
                # 详情
                detail_item = QTableWidgetItem(action)
                detail_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                self.log_table.setItem(row_idx, 3, detail_item)
            
        except Exception as e:
            QMessageBox.warning(
                self,
                "错误",
                f"加载日志失败：\n{str(e)}"
            )
    
    def _get_action_type(self, action):
        """根据action内容判断操作类型"""
        action_lower = action.lower()
        if '登录' in action or 'login' in action_lower:
            return "🔐 登录"
        elif '退出' in action or 'logout' in action_lower:
            return "🚪 退出"
        elif '新增' in action or '添加' in action or 'insert' in action_lower:
            return "➕ 新增"
        elif '删除' in action or 'delete' in action_lower:
            return "❌ 删除"
        elif '修改' in action or '更新' in action or 'update' in action_lower:
            return "✏️ 修改"
        elif '导入' in action or 'import' in action_lower:
            return "📥 导入"
        elif '导出' in action or 'export' in action_lower:
            return "📤 导出"
        elif '注册' in action or 'register' in action_lower:
            return "📝 注册"
        elif '规划' in action or '优化' in action:
            return "🎯 规划"
        elif '保存' in action or 'save' in action_lower:
            return "💾 保存"
        else:
            return "📌 其他"
    
    def clear_logs(self):
        """清空所有日志"""
        reply = QMessageBox.question(
            self,
            "确认清空",
            "确定要清空所有系统日志吗？\n此操作不可恢复！",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                conn = sqlite3.connect(DB_PATH)
                cursor = conn.cursor()
                cursor.execute("DELETE FROM logs")
                conn.commit()
                conn.close()
                
                self.load_logs()
                QMessageBox.information(self, "成功", "日志已清空")
                
            except Exception as e:
                QMessageBox.warning(
                    self,
                    "错误",
                    f"清空日志失败：\n{str(e)}"
                )


# ================= 测试运行 =================
if __name__ == "__main__":

    from PyQt6.QtWidgets import QApplication
    import sys

    app = QApplication(sys.argv)

    win = OrderManagerWindow()

    win.show()

    sys.exit(app.exec())