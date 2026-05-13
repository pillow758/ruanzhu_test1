import sys
import os
import sqlite3
import math
import random
from datetime import datetime
from database.db_manager import save_routes
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QHBoxLayout, QVBoxLayout,
    QPushButton, QLabel, QFrame, QTableWidget, QTableWidgetItem,
    QHeaderView, QInputDialog, QMessageBox, QStatusBar, QFileDialog,
    QLineEdit, QComboBox
)

try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
from PyQt6.QtCore import Qt, QTimer, QRectF, QPropertyAnimation, QEasingCurve, QPoint
from PyQt6.QtGui import QFont, QColor, QPalette, QLinearGradient, QPainter, QBrush, QPen
from PyQt6.QtWidgets import QGraphicsDropShadowEffect, QProgressBar, QSplitter, QScrollArea, QGridLayout
import pyqtgraph as pg
from echarts_window import EChartsAnalysis
from order_manager import OrderManagerWindow
from PyQt6.QtWidgets import QTabWidget
DB_PATH = os.path.join(os.path.dirname(__file__), "database/logistics.db")
DB_PATH = os.path.abspath(DB_PATH)
# 禁用OpenGL以避免黑屏问题
pg.setConfigOptions(useOpenGL=False)
pg.setConfigOptions(antialias=True)

# ====================== 参数配置 ======================
VEHICLE_CAPACITY = 3000
FIXED_COST = 400
COST_PER_KM = 2.5
COLORS = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', '#DDA0DD', '#98D8C8', '#F39C12', '#9B59B6', '#1ABC9C']

# ====================== 数据结构 ======================
class Node:
    def __init__(self, id, x, y, q):
        self.id = str(id)
        self.x = x
        self.y = y
        self.q = q

class Route:
    def __init__(self, nodes):
        self.nodes = nodes
        self.q = 0
        self.distance = 0
        self.cost = 0

# 全局数据（算法使用）
node_dict = {}
active_routes = []

# ====================== 核心算法 ======================
def calc_dist(n1, n2):
    return math.sqrt((n1.x - n2.x) ** 2 + (n1.y - n2.y) ** 2)

def evaluate_route(route_node_ids):
    total_q = 0
    total_dist = 0
    for nid in route_node_ids:
        if nid != 'DC':
            total_q += node_dict[nid].q
    for i in range(len(route_node_ids) - 1):
        n1 = node_dict[route_node_ids[i]]
        n2 = node_dict[route_node_ids[i + 1]]
        total_dist += calc_dist(n1, n2)
    cost = FIXED_COST + total_dist * COST_PER_KM if len(route_node_ids) > 2 else 0
    return total_q, total_dist, cost

def savings_algorithm(customer_nodes):
    """Clark-Wright 节约算法"""
    routes = [['DC', c.id, 'DC'] for c in customer_nodes]
    savings = []
    for i in range(len(customer_nodes)):
        for j in range(i + 1, len(customer_nodes)):
            n_i, n_j = customer_nodes[i], customer_nodes[j]
            d_dc_i = calc_dist(node_dict['DC'], n_i)
            d_dc_j = calc_dist(node_dict['DC'], n_j)
            d_i_j = calc_dist(n_i, n_j)
            sav = d_dc_i + d_dc_j - d_i_j
            if sav > 0:
                savings.append((n_i.id, n_j.id, sav))
    savings.sort(key=lambda x: x[2], reverse=True)

    for i_id, j_id, _ in savings:
        route_i_idx = route_j_idx = -1
        for idx, r in enumerate(routes):
            if r[1] == i_id and r[-2] == i_id:
                route_i_idx = idx
            elif r[-2] == i_id:
                route_i_idx = idx
            if r[1] == j_id and r[-2] == j_id:
                route_j_idx = idx
            elif r[1] == j_id:
                route_j_idx = idx
        if route_i_idx != -1 and route_j_idx != -1 and route_i_idx != route_j_idx:
            r_i, r_j = routes[route_i_idx], routes[route_j_idx]
            q_i, _, _ = evaluate_route(r_i)
            q_j, _, _ = evaluate_route(r_j)
            if q_i + q_j <= VEHICLE_CAPACITY:
                merged_route = r_i[:-1] + r_j[1:]
                routes.pop(max(route_i_idx, route_j_idx))
                routes.pop(min(route_i_idx, route_j_idx))
                routes.append(merged_route)

    final_routes = []
    for r in routes:
        q, dist, cost = evaluate_route(r)
        route_obj = Route(r)
        route_obj.q, route_obj.distance, route_obj.cost = q, dist, cost
        final_routes.append(route_obj)
    return final_routes

def trigger_new_order(new_node):
    """动态插入新订单"""
    global active_routes
    node_dict[new_node.id] = new_node
    best_route_idx = best_insert_pos = -1
    min_add_cost = float('inf')
    for r_idx, r in enumerate(active_routes):
        if r.q + new_node.q > VEHICLE_CAPACITY:
            continue
        for pos in range(1, len(r.nodes)):
            trial_nodes = r.nodes[:pos] + [new_node.id] + r.nodes[pos:]
            _, _, trial_cost = evaluate_route(trial_nodes)
            add_cost = trial_cost - r.cost
            if add_cost < min_add_cost:
                min_add_cost = add_cost
                best_route_idx = r_idx
                best_insert_pos = pos
    if best_route_idx != -1:
        r = active_routes[best_route_idx]
        r.nodes.insert(best_insert_pos, new_node.id)
        r.q, r.distance, r.cost = evaluate_route(r.nodes)
    else:
        r_new = ['DC', new_node.id, 'DC']
        q, dist, cost = evaluate_route(r_new)
        route_obj = Route(r_new)
        route_obj.q, route_obj.distance, route_obj.cost = q, dist, cost
        active_routes.append(route_obj)

def trigger_cancel_order(cancel_id):
    """取消订单"""
    global active_routes
    target_r_idx = next((i for i, r in enumerate(active_routes) if cancel_id in r.nodes), -1)
    if target_r_idx == -1:
        return
    r = active_routes[target_r_idx]
    r.nodes.remove(cancel_id)
    r.q, r.distance, r.cost = evaluate_route(r.nodes)
    if len(r.nodes) <= 2:
        active_routes.pop(target_r_idx)

def trigger_change_address(target_id, new_x, new_y):
    """变更地址（模拟先删除再插入）"""
    global active_routes
    target_node = node_dict[target_id]
    temp_q = target_node.q
    trigger_cancel_order(target_id)
    target_node.x, target_node.y, target_node.q = new_x, new_y, temp_q
    trigger_new_order(target_node)

# ====================== 主界面 ======================
class LogisticsApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("🚚 动态物流调度系统 by pillow")
        self.setGeometry(100, 100, 1400, 800)
        self.setup_ui()
        self.apply_dark_theme()

        # 动画状态
        self.vehicle_pos = []
        self.vehicle_colors = []
        self.unload_status = {}
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_plot)

        # 初始化数据并开始动画
        self.init_data()
        self.start_animation()

    def setup_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QHBoxLayout(central)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(10)

        # ---------- 左侧控制面板 ----------
        left_panel = QFrame()
        left_panel.setFixedWidth(360)
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(10, 10, 10, 10)
        left_layout.setSpacing(8)

        # 标题
        title = QLabel("📋 调度控制系统")
        title.setFont(QFont("Microsoft YaHei", 14, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        left_layout.addWidget(title)

        # 按钮组 - 分为两行布局
        # 第一行：主要操作按钮
        row1_layout = QHBoxLayout()
        row1_buttons = [
            ("🎯 静态规划", self.init_routes, "#27AE60"),
            ("➕ 新增", self.add_order, "#3498DB"),
            ("❌ 取消", self.cancel_order, "#E74C3C"),
        ]
        for text, slot, color in row1_buttons:
            btn = QPushButton(text)
            btn.setFont(QFont("Microsoft YaHei", 9))
            btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {color};
                    color: white;
                    border-radius: 6px;
                    padding: 8px;
                }}
                QPushButton:hover {{
                    background-color: {self.lighten_color(color)};
                }}
                QPushButton:pressed {{
                    background-color: {self.darken_color(color)};
                }}
            """)
            btn.clicked.connect(slot)
            row1_layout.addWidget(btn)
        left_layout.addLayout(row1_layout)
        
        # 第二行：辅助操作按钮
        row2_layout = QHBoxLayout()
        row2_buttons = [
            ("📍 变更", self.change_address, "#F39C12"),
            ("📥 导入Excel", self.import_excel_data, "#8b5cf6"),
        ]
        for text, slot, color in row2_buttons:
            btn = QPushButton(text)
            btn.setFont(QFont("Microsoft YaHei", 9))
            btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {color};
                    color: white;
                    border-radius: 6px;
                    padding: 8px;
                }}
                QPushButton:hover {{
                    background-color: {self.lighten_color(color)};
                }}
                QPushButton:pressed {{
                    background-color: {self.darken_color(color)};
                }}
            """)
            btn.clicked.connect(slot)
            row2_layout.addWidget(btn)
        left_layout.addLayout(row2_layout)

        # 统计信息
        stats_frame = QFrame()
        stats_frame.setStyleSheet("background-color: #2a2a3e; border-radius: 8px; padding: 10px;")
        stats_layout = QVBoxLayout(stats_frame)
        stats_layout.setSpacing(5)

        stats_title = QLabel("📊 运营统计")
        stats_title.setFont(QFont("Microsoft YaHei", 11, QFont.Weight.Bold))
        stats_layout.addWidget(stats_title)

        self.lbl_vehicle = QLabel("🚛 使用车辆: 0")
        self.lbl_load = QLabel("📦 总载重: 0 kg")
        self.lbl_dist = QLabel("📏 总距离: 0 km")
        self.lbl_cost = QLabel("💰 总成本: ¥ 0")
        for lbl in (self.lbl_vehicle, self.lbl_load, self.lbl_dist, self.lbl_cost):
            lbl.setFont(QFont("Microsoft YaHei", 9))
            stats_layout.addWidget(lbl)
        left_layout.addWidget(stats_frame)

        # 车辆详情表格
        table_header_layout = QHBoxLayout()
        table_label = QLabel("🚚 车辆配送详情")
        table_label.setFont(QFont("Microsoft YaHei", 11, QFont.Weight.Bold))
        table_header_layout.addWidget(table_label)
        table_header_layout.addStretch()
        
        # 全屏查看按钮
        self.fullscreen_table_btn = QPushButton("🔍 全屏查看")
        self.fullscreen_table_btn.setFont(QFont("Microsoft YaHei", 9))
        self.fullscreen_table_btn.setStyleSheet("""
            QPushButton {
                background-color: #3b82f6;
                color: white;
                border-radius: 6px;
                padding: 6px 12px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #60a5fa;
            }
            QPushButton:pressed {
                background-color: #2563eb;
            }
        """)
        self.fullscreen_table_btn.clicked.connect(self.show_fullscreen_table)
        table_header_layout.addWidget(self.fullscreen_table_btn)
        left_layout.addLayout(table_header_layout)

        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["车辆", "载重(kg)", "距离(km)", "成本(¥)", "路径"])
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setStyleSheet("""
            QTableWidget {
                background-color: #2a2a3e;
                gridline-color: #444;
                color: #e0e0e0;
                font: 9pt "Microsoft YaHei";
            }
            QHeaderView::section {
                background-color: #1e1e2e;
                color: #ccc;
                padding: 4px;
                border: 1px solid #444;
            }
        """)
        left_layout.addWidget(self.table, 1)

        main_layout.addWidget(left_panel)

        # ---------- 右侧动画面板 ----------
        right_panel = QFrame()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)

        # 新建 Tab
        self.tabs = QTabWidget()
        right_layout.addWidget(self.tabs)

        # 原来的动画页 - 使用 QWidget 作为容器
        self.dispatch_container = QWidget()
        dispatch_layout = QVBoxLayout(self.dispatch_container)
        dispatch_layout.setContentsMargins(0, 0, 0, 0)
        dispatch_layout.setSpacing(0)
        
        self.plot_widget = pg.PlotWidget()
        self.plot_widget.setBackground('#1e1e2e')
        self.plot_widget.setLabel('left', 'Y 坐标 (km)', color='#ccc')
        self.plot_widget.setLabel('bottom', 'X 坐标 (km)', color='#ccc')
        self.plot_widget.showGrid(x=True, y=True, alpha=0.3)
        self.plot_widget.setRange(xRange=[-2, 42], yRange=[-2, 42])
        self.plot_widget.setMinimumSize(400, 300)
        dispatch_layout.addWidget(self.plot_widget, 1)
        
        self.tabs.addTab(self.dispatch_container, "调度动画")

        # 数据分析页
        self.analysis_widget = EChartsAnalysis()
        self.tabs.addTab(self.analysis_widget, "数据分析")
        
        # 订单管理页
        self.order_manager_widget = OrderManagerWindow()
        self.tabs.addTab(self.order_manager_widget, "📦 订单管理")
        
        # 系统日志页
        self.log_widget = self.create_log_widget()
        self.tabs.addTab(self.log_widget, "📋 系统日志")
        
        # Tab切换时强制重绘
        self.tabs.currentChanged.connect(self.on_tab_changed)
        main_layout.addWidget(right_panel, 1)

        # ---------- 状态栏 ----------
        self.status_bar = QStatusBar()
        self.status_bar.setStyleSheet("background-color: #1e1e2e; color: #aaa;")
        self.status_msg = QLabel("✅ 系统就绪")
        self.status_msg.setFont(QFont("Microsoft YaHei", 9))
        self.status_bar.addWidget(self.status_msg)
        self.setStatusBar(self.status_bar)

    def on_tab_changed(self, index):
        """Tab切换时处理"""
        if index == 0:  # 调度动画页
            # 强制重绘plot_widget
            self.plot_widget.update()
            # 重新设置范围确保显示正确
            self.plot_widget.setRange(xRange=[-2, 42], yRange=[-2, 42])
        elif index == 4:  # 系统日志页
            # 刷新日志
            self.refresh_logs()
    
    def create_log_widget(self):
        """创建系统日志页面"""
        log_widget = QWidget()
        layout = QVBoxLayout(log_widget)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(10)
        
        # 标题
        title = QLabel("📋 系统操作日志")
        title.setFont(QFont("Microsoft YaHei", 16, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("""
            color: white;
            padding: 10px;
            background-color: #1e293b;
            border-radius: 10px;
        """)
        layout.addWidget(title)
        
        # 日志表格
        self.log_table = QTableWidget()
        self.log_table.setColumnCount(4)
        self.log_table.setHorizontalHeaderLabels(["时间", "用户", "操作类型", "详情"])
        self.log_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.log_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.log_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.log_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.log_table.setStyleSheet("""
            QTableWidget{
                background-color:#0f172a;
                color:white;
                gridline-color:#334155;
                font-size:13px;
            }
            QHeaderView::section{
                background-color:#1e293b;
                color:white;
                padding:8px;
                border:none;
                font-weight:bold;
            }
            QTableWidget::item:selected{
                background-color:#2563eb;
            }
        """)
        layout.addWidget(self.log_table)
        
        # 按钮区域
        btn_layout = QHBoxLayout()
        
        refresh_btn = QPushButton("🔄 刷新")
        refresh_btn.clicked.connect(self.refresh_logs)
        refresh_btn.setStyleSheet("""
            QPushButton{
                background-color:#3b82f6;
                color:white;
                padding:8px 16px;
                border-radius:8px;
                font-size:14px;
            }
            QPushButton:hover{
                background-color:#60a5fa;
            }
        """)
        btn_layout.addWidget(refresh_btn)
        
        clear_btn = QPushButton("🗑️ 清空日志")
        clear_btn.clicked.connect(self.clear_logs)
        clear_btn.setStyleSheet("""
            QPushButton{
                background-color:#dc2626;
                color:white;
                padding:8px 16px;
                border-radius:8px;
                font-size:14px;
            }
            QPushButton:hover{
                background-color:#ef4444;
            }
        """)
        btn_layout.addWidget(clear_btn)
        
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # 初始加载日志
        self.refresh_logs()
        
        return log_widget
    
    def refresh_logs(self):
        """刷新系统日志"""
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            
            # 获取最近的100条日志
            cursor.execute("""
                SELECT timestamp, username, action
                FROM logs
                ORDER BY timestamp DESC
                LIMIT 100
            """)
            
            logs = cursor.fetchall()
            conn.close()
            
            self.log_table.setRowCount(len(logs))
            
            for row_idx, (timestamp, username, action) in enumerate(logs):
                # 时间
                time_item = QTableWidgetItem(str(timestamp))
                time_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.log_table.setItem(row_idx, 0, time_item)
                
                # 用户
                user_item = QTableWidgetItem(str(username))
                user_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.log_table.setItem(row_idx, 1, user_item)
                
                # 操作类型
                action_type = self._get_action_type(action)
                type_item = QTableWidgetItem(action_type)
                type_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.log_table.setItem(row_idx, 2, type_item)
                
                # 详情
                detail_item = QTableWidgetItem(action)
                detail_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                self.log_table.setItem(row_idx, 3, detail_item)
            
        except Exception as e:
            print(f"刷新日志失败: {e}")
    
    def _get_action_type(self, action):
        """根据action内容判断操作类型"""
        if '登录' in action or 'login' in action.lower():
            return "🔐 登录"
        elif '退出' in action or 'logout' in action.lower():
            return "🚪 退出"
        elif '新增' in action or '添加' in action or 'insert' in action.lower():
            return "➕ 新增"
        elif '删除' in action or 'delete' in action.lower():
            return "❌ 删除"
        elif '修改' in action or '更新' in action or 'update' in action.lower():
            return "✏️ 修改"
        elif '导入' in action or 'import' in action.lower():
            return "📥 导入"
        elif '导出' in action or 'export' in action.lower():
            return "📤 导出"
        elif '注册' in action or 'register' in action.lower():
            return "📝 注册"
        elif '规划' in action or '优化' in action:
            return "🎯 规划"
        elif '保存' in action or 'save' in action.lower():
            return "💾 保存"
        else:
            return "📌 其他"
    
    def clear_logs(self):
        """清空所有日志"""
        reply = QMessageBox.question(
            self,
            "确认清空",
            "确定要清空所有系统日志吗？\n此操作不可恢复！",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                conn = sqlite3.connect(DB_PATH)
                cursor = conn.cursor()
                cursor.execute("DELETE FROM logs")
                conn.commit()
                conn.close()
                
                self.refresh_logs()
                QMessageBox.information(self, "成功", "日志已清空")
                
            except Exception as e:
                QMessageBox.warning(
                    self,
                    "错误",
                    f"清空日志失败：\n{str(e)}"
                )

    def apply_dark_theme(self):
        """全局深色科技风样式"""
        self.setStyleSheet("""
            QMainWindow {
                background-color: #12121a;
            }
            QWidget {
                background-color: #1e1e2e;
                color: #cdd6f4;
                font-family: "Microsoft YaHei";
            }
            QFrame {
                background-color: #1e1e2e;
                border: 1px solid #333;
                border-radius: 10px;
            }
            QLabel {
                background-color: transparent;
            }
        """)

    def lighten_color(self, hex_color):
        """简易颜色变亮"""
        r, g, b = int(hex_color[1:3], 16), int(hex_color[3:5], 16), int(hex_color[5:7], 16)
        r = min(255, r + 30); g = min(255, g + 30); b = min(255, b + 30)
        return f"#{r:02x}{g:02x}{b:02x}"

    def darken_color(self, hex_color):
        r, g, b = int(hex_color[1:3], 16), int(hex_color[3:5], 16), int(hex_color[5:7], 16)
        r = max(0, r - 30); g = max(0, g - 30); b = max(0, b - 30)
        return f"#{r:02x}{g:02x}{b:02x}"

    def init_data(self):
        """初始化示例数据"""
        global node_dict, active_routes
        node_dict.clear()
        active_routes.clear()
        node_dict['DC'] = Node('DC', 20, 20, 0)
        raw_customers = [
            Node('A', 25, 15, 800), Node('B', 15, 10, 600), Node('C', 30, 25, 700),
            Node('D', 10, 30, 500), Node('E', 28, 8, 900), Node('F', 5, 20, 400),
        ]
        for c in raw_customers:
            node_dict[c.id] = c
        self.init_routes()

    def init_routes(self):
        """静态优化路线"""
        global active_routes
        customers = [node_dict[nid] for nid in node_dict if nid != 'DC']
        active_routes = savings_algorithm(customers)
        self.save_routes_to_db() 
        self.reset_animation()
        self.update_stats()
        self.status_msg.setText("✅ 静态规划完成，路线已优化")
        # 延迟显示完成消息，避免影响初始化渲染
        QTimer.singleShot(100, lambda: QMessageBox.information(self, "完成",
            f"静态规划完成！\n使用车辆: {len(active_routes)} 辆\n总成本: ¥{sum(r.cost for r in active_routes):.1f}"))
    
    def save_routes_to_db(self):
        """把 active_routes 写入 routes 表供 驾驶员 使用"""
        global active_routes

        # ======================
        # 调试输出 1：看看有没有进这个函数
        # ======================
        print("=" * 60)
        print("进入函数：save_routes_to_db()")
        print("当前路线数量：", len(active_routes) if active_routes else 0)

        if not active_routes:
            print("错误：没有路线，不保存到数据库！")
            print("=" * 60)
            return

        # ======================
        # 调试输出 2：打印要保存的所有路线数据
        # ======================
        routes_list = []
        create_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        for idx, r in enumerate(active_routes):
            vehicle_id = f"车辆{idx+1}"
            q = r.q
            distance = r.distance
            cost = r.cost
            nodes_str = " → ".join(r.nodes)
            routes_list.append((
                vehicle_id,
                q,
                distance,
                cost,
                nodes_str,
                create_time,
                "",
                "",
                0,
                0
            ))

            print(f"要保存的车辆 {idx+1}：")
            print(f"       ID: {vehicle_id}")
            print(f"       载重: {q}")
            print(f"       距离: {distance}")
            print(f"       成本: {cost}")
            print(f"       路线: {nodes_str}")
            print(f"       创建时间: {create_time}")
            print("-" * 40)

        # ======================
        # 调试输出 3：开始保存
        # ======================
        try:
            from database.db_manager import save_routes
            print("成功导入 save_routes 函数")

            save_routes(routes_list)

            #print("✅ 路线已经成功保存到数据库！")
            
            from database.db_manager import save_routes, save_customers_nodes
            save_customers_nodes(node_dict)
            #print("✅ 节点已经成功保存到数据库！")
        except Exception as e:
            print("❌ 保存失败，错误信息：", str(e))

        print("=" * 60)
   
    def reset_animation(self):
        """重置动画状态"""
        self.vehicle_pos = [0.0] * len(active_routes)
        self.vehicle_colors = [COLORS[i % len(COLORS)] for i in range(len(active_routes))]
        self.unload_status = {i: False for i in range(len(active_routes))}

    def start_animation(self):
        # 先执行一次绘图，确保初始显示
        self.update_plot()
        self.timer.start(100)  # 10帧/秒，降低频率提高性能

    def draw_road_background(self):
        """绘制模拟道路背景"""
        # 绘制主要道路（横向和纵向）
        road_pen = pg.mkPen(color='#2a3f5f', width=3)
        road_pen_light = pg.mkPen(color='#1e3a5f', width=2)
        
        # 主要横向道路
        main_roads_y = [5, 10, 15, 20, 25, 30, 35]
        for y in main_roads_y:
            line = pg.PlotDataItem([0, 40], [y, y], pen=road_pen)
            self.plot_widget.addItem(line)
        
        # 主要纵向道路
        main_roads_x = [5, 10, 15, 20, 25, 30, 35]
        for x in main_roads_x:
            line = pg.PlotDataItem([x, x], [0, 40], pen=road_pen)
            self.plot_widget.addItem(line)
        
        # 次要道路（虚线效果）
        minor_roads_y = [2.5, 7.5, 12.5, 17.5, 22.5, 27.5, 32.5, 37.5]
        for y in minor_roads_y:
            line = pg.PlotDataItem([0, 40], [y, y], pen=road_pen_light)
            self.plot_widget.addItem(line)
        
        minor_roads_x = [2.5, 7.5, 12.5, 17.5, 22.5, 27.5, 32.5, 37.5]
        for x in minor_roads_x:
            line = pg.PlotDataItem([x, x], [0, 40], pen=road_pen_light)
            self.plot_widget.addItem(line)
        
        # 绘制一些"建筑物"区块（深色矩形区域）
        building_brush = pg.mkBrush('#151d2e')
        building_pen = pg.mkPen(color='#0f1720', width=1)
        
        # 添加一些随机的建筑物区块
        buildings = [
            (2, 2, 3, 3), (7, 2, 3, 3), (12, 2, 3, 3), (17, 2, 3, 3),
            (22, 2, 3, 3), (27, 2, 3, 3), (32, 2, 3, 3), (37, 2, 2.5, 3),
            (2, 7, 3, 3), (7, 7, 3, 3), (12, 7, 3, 3), (17, 7, 3, 3),
            (22, 7, 3, 3), (27, 7, 3, 3), (32, 7, 3, 3), (37, 7, 2.5, 3),
            (2, 12, 3, 3), (7, 12, 3, 3), (12, 12, 3, 3), (17, 12, 3, 3),
            (22, 12, 3, 3), (27, 12, 3, 3), (32, 12, 3, 3), (37, 12, 2.5, 3),
            (2, 17, 3, 3), (7, 17, 3, 3), (12, 17, 3, 3), (17, 17, 3, 3),
            (22, 17, 3, 3), (27, 17, 3, 3), (32, 17, 3, 3), (37, 17, 2.5, 3),
            (2, 22, 3, 3), (7, 22, 3, 3), (12, 22, 3, 3), (17, 22, 3, 3),
            (22, 22, 3, 3), (27, 22, 3, 3), (32, 22, 3, 3), (37, 22, 2.5, 3),
            (2, 27, 3, 3), (7, 27, 3, 3), (12, 27, 3, 3), (17, 27, 3, 3),
            (22, 27, 3, 3), (27, 27, 3, 3), (32, 27, 3, 3), (37, 27, 2.5, 3),
            (2, 32, 3, 3), (7, 32, 3, 3), (12, 32, 3, 3), (17, 32, 3, 3),
            (22, 32, 3, 3), (27, 32, 3, 3), (32, 32, 3, 3), (37, 32, 2.5, 3),
            (2, 37, 3, 2.5), (7, 37, 3, 2.5), (12, 37, 3, 2.5), (17, 37, 3, 2.5),
            (22, 37, 3, 2.5), (27, 37, 3, 2.5), (32, 37, 3, 2.5), (37, 37, 2.5, 2.5),
        ]
        
        for bx, by, bw, bh in buildings:
            # 使用散点图模拟矩形区域
            building = pg.ScatterPlotItem(
                [bx + bw/2], [by + bh/2], 
                symbol='s', size=bw*20, 
                pen=building_pen, brush=building_brush
            )
            self.plot_widget.addItem(building)

    def update_plot(self):
        """每帧重绘全部内容（高性能优化版）"""
        self.plot_widget.clear()
        
        # 绘制道路背景（仅在节点数较少时绘制详细背景）
        if len(node_dict) <= 50:
            self.draw_road_background()
        else:
            # 大量数据时简化背景
            self.draw_simple_background()
        
        # 绘制配送中心
        dc = node_dict['DC']
        dc_item = pg.ScatterPlotItem(
            [dc.x], [dc.y], symbol='s', size=20,
            pen=pg.mkPen(color='#FFD700', width=3), brush=pg.mkBrush('#2C3E50')
        )
        self.plot_widget.addItem(dc_item)
        dc_text = pg.TextItem('DC', color='#FFD700', anchor=(0.5, -0.5))
        dc_text.setFont(QFont('Microsoft YaHei', 10, QFont.Weight.Bold))
        dc_text.setPos(dc.x, dc.y + 0.8)
        self.plot_widget.addItem(dc_text)

        # 绘制所有客户点（批量渲染优化）
        client_nodes = [(n.x, n.y) for nid, n in node_dict.items() if nid != 'DC']
        if client_nodes:
            client_x = [n[0] for n in client_nodes]
            client_y = [n[1] for n in client_nodes]
            clients = pg.ScatterPlotItem(
                client_x, client_y, symbol='o', size=8,
                pen=pg.mkPen(color='white', width=1), brush=pg.mkBrush('#3498DB')
            )
            self.plot_widget.addItem(clients)
            
            # 节点数少时显示标签，多时只显示部分
            if len(client_nodes) <= 30:
                for nid, n in node_dict.items():
                    if nid == 'DC':
                        continue
                    label = pg.TextItem(f'{nid}', color='#b0b0b0', anchor=(0.5, -0.5))
                    label.setFont(QFont('Microsoft YaHei', 7))
                    label.setPos(n.x, n.y + 0.3)
                    self.plot_widget.addItem(label)

        # 绘制路线与车辆动画（性能优化）
        # 当路线过多时，只绘制部分车辆的动画
        max_display_routes = min(len(active_routes), 20 if len(active_routes) > 50 else len(active_routes))
        
        for idx in range(max_display_routes):
            r = active_routes[idx]
            xs = [node_dict[nid].x for nid in r.nodes]
            ys = [node_dict[nid].y for nid in r.nodes]
            color = self.vehicle_colors[idx]

            # 路线（简化渲染）
            line_pen = pg.mkPen(color, width=1.5 if len(active_routes) > 30 else 2, 
                               style=Qt.PenStyle.SolidLine if len(active_routes) > 30 else Qt.PenStyle.DashLine)
            line = pg.PlotDataItem(xs, ys, pen=line_pen, symbol=None)
            self.plot_widget.addItem(line)

            # 车辆动画位置（只显示部分车辆）
            if idx < len(self.vehicle_pos):
                pos = self.vehicle_pos[idx]
                if pos >= len(r.nodes) - 1:
                    self.vehicle_pos[idx] = 0.0
                    pos = 0.0
                i = int(pos)
                j = min(i + 1, len(r.nodes) - 1)
                t = pos - i
                x = xs[i] + (xs[j] - xs[i]) * t
                y = ys[i] + (ys[j] - ys[i]) * t

                # 简化车辆渲染
                vehicle = pg.ScatterPlotItem(
                    [x], [y], symbol='o', size=10,
                    pen=pg.mkPen('white', width=1), brush=pg.mkBrush(color)
                )
                self.plot_widget.addItem(vehicle)

                self.vehicle_pos[idx] += 0.05  # 增加移动步长，减少重绘次数

    def draw_simple_background(self):
        """简化背景绘制（用于大量数据时）"""
        # 只绘制主要道路
        road_pen = pg.mkPen(color='#2a3f5f', width=2)
        for i in range(0, 41, 10):
            line_h = pg.PlotDataItem([0, 40], [i, i], pen=road_pen)
            line_v = pg.PlotDataItem([i, i], [0, 40], pen=road_pen)
            self.plot_widget.addItem(line_h)
            self.plot_widget.addItem(line_v)

    def update_stats(self):
        """刷新统计信息与表格"""
        total_dist = sum(r.distance for r in active_routes)
        total_cost = sum(r.cost for r in active_routes)
        total_q = sum(r.q for r in active_routes)
        self.lbl_vehicle.setText(f"🚛 使用车辆: {len(active_routes)}")
        self.lbl_load.setText(f"📦 总载重: {total_q:.0f} kg")
        self.lbl_dist.setText(f"📏 总距离: {total_dist:.1f} km")
        self.lbl_cost.setText(f"💰 总成本: ¥ {total_cost:.1f}")

        # 更新表格（大量数据时限制显示行数）
        max_table_rows = min(len(active_routes), 50)  # 最多显示50行
        self.table.setRowCount(max_table_rows)
        for idx in range(max_table_rows):
            r = active_routes[idx]
            path_str = " → ".join(r.nodes[:3])
            if len(r.nodes) > 4:
                path_str += " …"
            items = [
                QTableWidgetItem(f"车辆{idx+1}"),
                QTableWidgetItem(f"{r.q:.0f}"),
                QTableWidgetItem(f"{r.distance:.1f}"),
                QTableWidgetItem(f"¥{r.cost:.1f}"),
                QTableWidgetItem(path_str),
            ]
            for col, item in enumerate(items):
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(idx, col, item)

        self.status_msg.setText(
            f"📅 最后更新 | 📊 车辆: {len(active_routes)} | 成本: ¥{total_cost:.0f}"
        )

    # ---------- 按钮槽函数 ----------
    def add_order(self):
        nid, ok = QInputDialog.getText(self, "新增订单", "请输入订单ID:")
        if not ok or not nid.strip():
            return
        nid = nid.strip()
        if nid in node_dict:
            QMessageBox.warning(self, "错误", "订单ID已存在！")
            return
        x, ok = QInputDialog.getDouble(self, "新增订单", "X坐标 (0-40):", 20, 0, 40, 1)
        if not ok: return
        y, ok = QInputDialog.getDouble(self, "新增订单", "Y坐标 (0-40):", 20, 0, 40, 1)
        if not ok: return
        q, ok = QInputDialog.getDouble(self, "新增订单", "重量 (kg):", 500, 1, 3000, 1)
        if not ok: return

        new_node = Node(nid, x, y, q)
        global active_routes
        trigger_new_order(new_node)
        self.save_routes_to_db()
        self.reset_animation()
        self.update_stats()
        self.status_msg.setText(f"✅ 已新增订单 {nid}")

    def cancel_order(self):
        customers = [nid for nid in node_dict if nid != 'DC']
        if not customers:
            QMessageBox.warning(self, "警告", "没有可取消的订单！")
            return
        nid, ok = QInputDialog.getItem(self, "取消订单", "选择订单ID:", customers, 0, False)
        if ok and nid:
            global active_routes
            trigger_cancel_order(nid)
            self.save_routes_to_db()
            self.reset_animation()
            self.update_stats()
            self.status_msg.setText(f"✅ 已取消订单 {nid}")

    def change_address(self):
        customers = [nid for nid in node_dict if nid != 'DC']
        if not customers:
            QMessageBox.warning(self, "警告", "没有可修改的订单！")
            return
        nid, ok = QInputDialog.getItem(self, "地址变更", "选择订单:", customers, 0, False)
        if not ok: return
        x, ok = QInputDialog.getDouble(self, "地址变更", "新X坐标 (0-40):", 20, 0, 40, 1)
        if not ok: return
        y, ok = QInputDialog.getDouble(self, "地址变更", "新Y坐标 (0-40):", 20, 0, 40, 1)
        if not ok: return
        global active_routes
        trigger_change_address(nid, x, y)
        self.save_routes_to_db()
        self.reset_animation()
        self.update_stats()
        self.status_msg.setText(f"✅ 已更新订单 {nid} 地址")

    def show_fullscreen_table(self):
        """全屏显示车辆配送详情表格"""
        from PyQt6.QtWidgets import QDialog, QVBoxLayout
        
        dialog = QDialog(self)
        dialog.setWindowTitle("🚚 车辆配送详情 - 全屏查看")
        dialog.resize(1200, 800)
        
        layout = QVBoxLayout(dialog)
        
        # 创建全屏表格
        fullscreen_table = QTableWidget()
        fullscreen_table.setColumnCount(5)
        fullscreen_table.setHorizontalHeaderLabels(["车辆", "载重(kg)", "距离(km)", "成本(¥)", "路径"])
        fullscreen_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        fullscreen_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        fullscreen_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        fullscreen_table.setStyleSheet("""
            QTableWidget {
                background-color: #2a2a3e;
                gridline-color: #444;
                color: #e0e0e0;
                font: 10pt "Microsoft YaHei";
            }
            QHeaderView::section {
                background-color: #1e1e2e;
                color: #ccc;
                padding: 8px;
                border: 1px solid #444;
                font-weight: bold;
                font-size: 12px;
            }
            QTableWidget::item {
                padding: 6px;
            }
            QTableWidget::item:selected {
                background-color: #2563eb;
            }
        """)
        
        # 填充所有数据
        fullscreen_table.setRowCount(len(active_routes))
        for idx, r in enumerate(active_routes):
            path_str = " → ".join(r.nodes)
            items = [
                QTableWidgetItem(f"车辆{idx+1}"),
                QTableWidgetItem(f"{r.q:.0f}"),
                QTableWidgetItem(f"{r.distance:.1f}"),
                QTableWidgetItem(f"¥{r.cost:.1f}"),
                QTableWidgetItem(path_str),
            ]
            for col, item in enumerate(items):
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                fullscreen_table.setItem(idx, col, item)
        
        layout.addWidget(fullscreen_table)
        
        # 添加关闭按钮
        close_btn = QPushButton("✕ 关闭")
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc2626;
                color: white;
                border-radius: 6px;
                padding: 10px 20px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #ef4444;
            }
        """)
        close_btn.clicked.connect(dialog.close)
        layout.addWidget(close_btn)
        
        dialog.setStyleSheet("""
            QDialog {
                background-color: #1e1e2e;
            }
        """)
        
        dialog.exec()

    def import_excel_data(self):
        """从Excel导入订单数据 - 同步到订单管理系统"""
        if not EXCEL_AVAILABLE:
            QMessageBox.warning(
                self,
                "错误",
                "请先安装 openpyxl 库：\npip install openpyxl"
            )
            return
        
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "导入Excel订单数据",
            "",
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            global active_routes, node_dict
            imported_count = 0
            
            # 连接数据库同步到orders表
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            
            # 从第二行开始读取（跳过表头）
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    if len(row) < 5:
                        continue
                    
                    customer_name = str(row[1]) if row[1] else ""
                    x = float(row[2]) if row[2] else 0
                    y = float(row[3]) if row[3] else 0
                    demand = float(row[4]) if row[4] else 0
                    
                    if not customer_name or demand <= 0:
                        continue
                    
                    # 生成唯一ID
                    nid = f"C{customer_name}"
                    counter = 1
                    while nid in node_dict:
                        nid = f"C{customer_name}_{counter}"
                        counter += 1
                    
                    # 创建新节点并触发插入
                    new_node = Node(nid, x, y, demand)
                    node_dict[nid] = new_node
                    trigger_new_order(new_node)
                    
                    # 同步插入到 orders 表（订单管理系统使用）
                    create_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    cursor.execute("""
                        INSERT INTO orders (customer_name, x, y, demand, status, create_time)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (customer_name, x, y, demand, 'pending', create_time))
                    
                    imported_count += 1
                    
                except Exception as e:
                    print(f"导入行失败: {row}, 错误: {e}")
            
            conn.commit()
            conn.close()
            
            if imported_count > 0:
                self.save_routes_to_db()
                self.reset_animation()
                self.update_stats()
                self.status_msg.setText(f"✅ 成功导入 {imported_count} 条订单数据")
                QMessageBox.information(
                    self,
                    "完成",
                    f"成功导入 {imported_count} 条订单数据！\n系统已自动重新规划路线。\n订单管理中心数据已同步更新。"
                )
            else:
                QMessageBox.warning(
                    self,
                    "提示",
                    "没有成功导入任何订单数据，请检查Excel格式。"
                )
            
        except Exception as e:
            QMessageBox.warning(
                self,
                "错误",
                f"导入失败：\n{str(e)}"
            )

    def batch_simulation(self):
        num, ok = QInputDialog.getInt(self, "批量随机模拟", "生成订单数量:", 5, 1, 20, 1)
        if not ok: return
        self.reset_system()  # 先清空
        global active_routes
        new_ids = []
        for _ in range(num):
            nid = f"R{random.randint(100, 999)}"
            while nid in node_dict:
                nid = f"R{random.randint(100, 999)}"
            x = random.uniform(2, 38)
            y = random.uniform(2, 38)
            q = random.randint(200, 1500)
            new_node = Node(nid, x, y, q)
            trigger_new_order(new_node)
            new_ids.append(nid)
        self.reset_animation()
        self.save_routes_to_db()
        self.update_stats()
        self.status_msg.setText(f"✅ 批量模拟完成，新增 {num} 个随机订单")
        QMessageBox.information(self, "完成", f"已生成 {num} 个随机订单！\nID示例: {', '.join(new_ids[:5])}" +
                                ("…" if num > 5 else ""))

    def reset_system(self):
        reply = QMessageBox.question(self, "确认", "重置将清除所有订单数据，确定吗？",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.init_data()

# ====================== 现代化UI组件类 ======================

class ModernCard(QFrame):
    """现代化卡片组件"""
    def __init__(self, title="", icon="", parent=None):
        super().__init__(parent)
        self.setObjectName("modernCard")
        self.title = title
        self.icon = icon
        self.setup_ui()
        self.apply_shadow()
        
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # 标题栏
        if self.title:
            header = QHBoxLayout()
            if self.icon:
                icon_label = QLabel(self.icon)
                icon_label.setFont(QFont("Microsoft YaHei", 16))
                header.addWidget(icon_label)
            title_label = QLabel(self.title)
            title_label.setFont(QFont("Microsoft YaHei", 14, QFont.Weight.Bold))
            title_label.setStyleSheet("color: #cdd6f4;")
            header.addWidget(title_label)
            header.addStretch()
            layout.addLayout(header)
            
        self.content_layout = QVBoxLayout()
        layout.addLayout(self.content_layout)
        
        self.setStyleSheet("""
            #modernCard {
                background-color: #1e1e2e;
                border-radius: 16px;
                border: 1px solid #313244;
            }
            #modernCard:hover {
                border: 1px solid #585b70;
            }
        """)
        
    def apply_shadow(self):
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(20)
        shadow.setColor(QColor(0, 0, 0, 80))
        shadow.setOffset(0, 4)
        self.setGraphicsEffect(shadow)
        
    def add_widget(self, widget):
        self.content_layout.addWidget(widget)


class AnimatedButton(QPushButton):
    """带动画效果的按钮"""
    def __init__(self, text="", icon="", parent=None):
        super().__init__(f"{icon} {text}" if icon else text, parent)
        self.setFont(QFont("Microsoft YaHei", 11))
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setup_animation()
        
    def setup_animation(self):
        self.animation = QPropertyAnimation(self, b"geometry")
        self.animation.setDuration(150)
        self.animation.setEasingCurve(QEasingCurve.Type.OutQuad)
        
    def enterEvent(self, event):
        self.setStyleSheet(self.styleSheet() + "transform: scale(1.05);")
        super().enterEvent(event)
        
    def leaveEvent(self, event):
        super().leaveEvent(event)


class GradientProgressBar(QProgressBar):
    """渐变进度条"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setTextVisible(True)
        self.setFont(QFont("Microsoft YaHei", 9))
        self.setStyleSheet("""
            QProgressBar {
                border: none;
                border-radius: 10px;
                background-color: #313244;
                text-align: center;
                color: #cdd6f4;
                font-weight: bold;
            }
            QProgressBar::chunk {
                border-radius: 10px;
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 0,
                    stop: 0 #3b82f6,
                    stop: 0.5 #8b5cf6,
                    stop: 1 #ec4899
                );
            }
        """)


class StatBadge(QFrame):
    """统计徽章组件"""
    def __init__(self, label="", value="", unit="", color="#3b82f6", parent=None):
        super().__init__(parent)
        self.setFixedSize(140, 90)
        self.setStyleSheet(f"""
            StatBadge {{
                background-color: {color}20;
                border-radius: 12px;
                border: 2px solid {color}40;
            }}
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(4)
        
        self.label = QLabel(label)
        self.label.setFont(QFont("Microsoft YaHei", 10))
        self.label.setStyleSheet(f"color: {color};")
        self.label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.label)
        
        self.value = QLabel(value)
        self.value.setFont(QFont("Microsoft YaHei", 20, QFont.Weight.Bold))
        self.value.setStyleSheet(f"color: {color};")
        self.value.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.value)
        
        if unit:
            self.unit = QLabel(unit)
            self.unit.setFont(QFont("Microsoft YaHei", 9))
            self.unit.setStyleSheet(f"color: {color}80;")
            self.unit.setAlignment(Qt.AlignmentFlag.AlignCenter)
            layout.addWidget(self.unit)
        
    def set_value(self, value):
        self.value.setText(str(value))


class ModernTabBar(QFrame):
    """现代化标签栏"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedHeight(60)
        self.setStyleSheet("""
            ModernTabBar {
                background-color: #181825;
                border-radius: 12px;
                border: 1px solid #313244;
            }
        """)
        
        self.layout = QHBoxLayout(self)
        self.layout.setContentsMargins(10, 5, 10, 5)
        self.layout.setSpacing(8)
        self.buttons = []
        
    def add_tab(self, text, icon=""):
        btn = QPushButton(f"{icon} {text}" if icon else text)
        btn.setFont(QFont("Microsoft YaHei", 11))
        btn.setCheckable(True)
        btn.setCursor(Qt.CursorShape.PointingHandCursor)
        btn.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: #6c7086;
                border: none;
                border-radius: 8px;
                padding: 10px 20px;
                font-weight: 500;
            }
            QPushButton:hover {
                background-color: #313244;
                color: #cdd6f4;
            }
            QPushButton:checked {
                background-color: #3b82f6;
                color: white;
                font-weight: bold;
            }
        """)
        self.buttons.append(btn)
        self.layout.addWidget(btn)
        return btn


class SearchBox(QLineEdit):
    """现代化搜索框"""
    def __init__(self, placeholder="搜索...", parent=None):
        super().__init__(parent)
        self.setPlaceholderText(f"🔍 {placeholder}")
        self.setFont(QFont("Microsoft YaHei", 11))
        self.setMinimumHeight(40)
        self.setStyleSheet("""
            SearchBox {
                background-color: #181825;
                border: 2px solid #313244;
                border-radius: 12px;
                padding: 8px 16px;
                color: #cdd6f4;
            }
            SearchBox:focus {
                border: 2px solid #3b82f6;
                background-color: #1e1e2e;
            }
            SearchBox::placeholder {
                color: #6c7086;
            }
        """)


class NotificationBadge(QLabel):
    """通知徽章"""
    def __init__(self, count=0, parent=None):
        super().__init__(str(count) if count > 0 else "", parent)
        self.setFont(QFont("Microsoft YaHei", 9, QFont.Weight.Bold))
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setStyleSheet("""
            background-color: #ef4444;
            color: white;
            border-radius: 10px;
            padding: 2px 8px;
            min-width: 20px;
            min-height: 20px;
        """)
        self.setVisible(count > 0)
        
    def set_count(self, count):
        self.setText(str(count) if count > 0 else "")
        self.setVisible(count > 0)


class ToolButton(QPushButton):
    """工具按钮"""
    def __init__(self, icon="", tooltip="", parent=None):
        super().__init__(icon, parent)
        self.setToolTip(tooltip)
        self.setFixedSize(36, 36)
        self.setFont(QFont("Microsoft YaHei", 14))
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setStyleSheet("""
            ToolButton {
                background-color: #313244;
                border: none;
                border-radius: 8px;
                color: #cdd6f4;
            }
            ToolButton:hover {
                background-color: #585b70;
            }
            ToolButton:pressed {
                background-color: #6c7086;
            }
        """)


class StatusIndicator(QFrame):
    """状态指示器"""
    def __init__(self, status="normal", text="", parent=None):
        super().__init__(parent)
        self.setFixedHeight(32)
        
        colors = {
            "normal": ("#10b981", "#10b98120"),
            "warning": ("#f59e0b", "#f59e0b20"),
            "error": ("#ef4444", "#ef444420"),
            "info": ("#3b82f6", "#3b82f620")
        }
        
        color, bg = colors.get(status, colors["normal"])
        
        self.setStyleSheet(f"""
            StatusIndicator {{
                background-color: {bg};
                border-radius: 16px;
                border: 1px solid {color}40;
            }}
        """)
        
        layout = QHBoxLayout(self)
        layout.setContentsMargins(12, 4, 12, 4)
        
        dot = QLabel("●")
        dot.setFont(QFont("Microsoft YaHei", 8))
        dot.setStyleSheet(f"color: {color};")
        layout.addWidget(dot)
        
        label = QLabel(text)
        label.setFont(QFont("Microsoft YaHei", 10))
        label.setStyleSheet(f"color: {color};")
        layout.addWidget(label)


class Divider(QFrame):
    """分隔线"""
    def __init__(self, orientation="horizontal", parent=None):
        super().__init__(parent)
        if orientation == "horizontal":
            self.setFixedHeight(1)
        else:
            self.setFixedWidth(1)
        self.setStyleSheet("background-color: #313244;")


class Avatar(QLabel):
    """头像组件"""
    def __init__(self, text="", size=40, parent=None):
        super().__init__(text, parent)
        self.setFixedSize(size, size)
        self.setFont(QFont("Microsoft YaHei", 14, QFont.Weight.Bold))
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setStyleSheet(f"""
            background: qlineargradient(
                x1: 0, y1: 0, x2: 1, y2: 1,
                stop: 0 #3b82f6,
                stop: 1 #8b5cf6
            );
            color: white;
            border-radius: {size//2}px;
            border: 2px solid #585b70;
        """)


class LoadingSpinner(QLabel):
    """加载动画"""
    def __init__(self, parent=None):
        super().__init__("⏳ 加载中...", parent)
        self.setFont(QFont("Microsoft YaHei", 12))
        self.setStyleSheet("color: #6c7086;")
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.dots = ["⏳ 加载中", "⏳ 加载中.", "⏳ 加载中..", "⏳ 加载中..."]
        self.current = 0
        
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_text)
        
    def start(self):
        self.timer.start(500)
        self.show()
        
    def stop(self):
        self.timer.stop()
        self.hide()
        
    def update_text(self):
        self.current = (self.current + 1) % len(self.dots)
        self.setText(self.dots[self.current])


class EmptyState(QFrame):
    """空状态提示"""
    def __init__(self, icon="📭", title="暂无数据", subtitle="", parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setSpacing(16)
        
        icon_label = QLabel(icon)
        icon_label.setFont(QFont("Microsoft YaHei", 48))
        icon_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(icon_label)
        
        title_label = QLabel(title)
        title_label.setFont(QFont("Microsoft YaHei", 16, QFont.Weight.Bold))
        title_label.setStyleSheet("color: #cdd6f4;")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        if subtitle:
            subtitle_label = QLabel(subtitle)
            subtitle_label.setFont(QFont("Microsoft YaHei", 12))
            subtitle_label.setStyleSheet("color: #6c7086;")
            subtitle_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            layout.addWidget(subtitle_label)


class Chip(QFrame):
    """标签芯片"""
    def __init__(self, text="", color="#3b82f6", removable=False, parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"""
            Chip {{
                background-color: {color}20;
                border-radius: 12px;
                border: 1px solid {color}40;
            }}
        """)
        
        layout = QHBoxLayout(self)
        layout.setContentsMargins(10, 4, 10, 4)
        layout.setSpacing(4)
        
        label = QLabel(text)
        label.setFont(QFont("Microsoft YaHei", 10))
        label.setStyleSheet(f"color: {color};")
        layout.addWidget(label)
        
        if removable:
            remove_btn = QLabel("✕")
            remove_btn.setFont(QFont("Microsoft YaHei", 10))
            remove_btn.setStyleSheet(f"color: {color}; cursor: pointer;")
            layout.addWidget(remove_btn)


class TimelineItem(QFrame):
    """时间线项"""
    def __init__(self, time="", title="", description="", active=True, parent=None):
        super().__init__(parent)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 8, 0, 8)
        layout.setSpacing(16)
        
        # 时间点
        dot_frame = QFrame()
        dot_frame.setFixedWidth(24)
        dot_layout = QVBoxLayout(dot_frame)
        dot_layout.setContentsMargins(0, 0, 0, 0)
        
        dot = QLabel("●")
        dot.setFont(QFont("Microsoft YaHei", 12))
        color = "#3b82f6" if active else "#6c7086"
        dot.setStyleSheet(f"color: {color};")
        dot.setAlignment(Qt.AlignmentFlag.AlignCenter)
        dot_layout.addWidget(dot)
        
        layout.addWidget(dot_frame)
        
        # 内容
        content = QVBoxLayout()
        content.setSpacing(4)
        
        time_label = QLabel(time)
        time_label.setFont(QFont("Microsoft YaHei", 10))
        time_label.setStyleSheet("color: #6c7086;")
        content.addWidget(time_label)
        
        title_label = QLabel(title)
        title_label.setFont(QFont("Microsoft YaHei", 12, QFont.Weight.Bold))
        title_label.setStyleSheet(f"color: {'#cdd6f4' if active else '#6c7086'};")
        content.addWidget(title_label)
        
        if description:
            desc_label = QLabel(description)
            desc_label.setFont(QFont("Microsoft YaHei", 10))
            desc_label.setStyleSheet("color: #6c7086;")
            desc_label.setWordWrap(True)
            content.addWidget(desc_label)
        
        layout.addLayout(content, 1)


class MetricCard(QFrame):
    """指标卡片"""
    def __init__(self, label="", value="", change="", positive=True, parent=None):
        super().__init__(parent)
        self.setFixedHeight(100)
        self.setStyleSheet("""
            MetricCard {
                background-color: #181825;
                border-radius: 12px;
                border: 1px solid #313244;
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(8)
        
        self.label = QLabel(label)
        self.label.setFont(QFont("Microsoft YaHei", 11))
        self.label.setStyleSheet("color: #6c7086;")
        layout.addWidget(self.label)
        
        value_layout = QHBoxLayout()
        
        self.value = QLabel(value)
        self.value.setFont(QFont("Microsoft YaHei", 24, QFont.Weight.Bold))
        self.value.setStyleSheet("color: #cdd6f4;")
        value_layout.addWidget(self.value)
        
        if change:
            self.change = QLabel(change)
            self.change.setFont(QFont("Microsoft YaHei", 11))
            color = "#10b981" if positive else "#ef4444"
            self.change.setStyleSheet(f"color: {color};")
            value_layout.addWidget(self.change)
            value_layout.addStretch()
        
        layout.addLayout(value_layout)


class SidebarItem(QPushButton):
    """侧边栏菜单项"""
    def __init__(self, icon="", text="", active=False, parent=None):
        super().__init__(f"{icon}  {text}" if icon else text, parent)
        self.setFont(QFont("Microsoft YaHei", 12))
        self.setCheckable(True)
        self.setChecked(active)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setStyleSheet("""
            SidebarItem {
                background-color: transparent;
                color: #6c7086;
                border: none;
                border-radius: 10px;
                padding: 12px 16px;
                text-align: left;
            }
            SidebarItem:hover {
                background-color: #313244;
                color: #cdd6f4;
            }
            SidebarItem:checked {
                background-color: #3b82f6;
                color: white;
                font-weight: bold;
            }
        """)


class FloatingActionButton(QPushButton):
    """浮动操作按钮"""
    def __init__(self, icon="+", parent=None):
        super().__init__(icon, parent)
        self.setFixedSize(56, 56)
        self.setFont(QFont("Microsoft YaHei", 20))
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setStyleSheet("""
            FloatingActionButton {
                background-color: #3b82f6;
                color: white;
                border: none;
                border-radius: 28px;
            }
            FloatingActionButton:hover {
                background-color: #60a5fa;
            }
            FloatingActionButton:pressed {
                background-color: #2563eb;
            }
        """)
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(20)
        shadow.setColor(QColor(59, 130, 246, 100))
        shadow.setOffset(0, 4)
        self.setGraphicsEffect(shadow)


class Breadcrumb(QFrame):
    """面包屑导航"""
    def __init__(self, items=None, parent=None):
        super().__init__(parent)
        items = items or []
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)
        
        for i, item in enumerate(items):
            if i > 0:
                sep = QLabel("›")
                sep.setFont(QFont("Microsoft YaHei", 14))
                sep.setStyleSheet("color: #6c7086;")
                layout.addWidget(sep)
            
            label = QLabel(item)
            label.setFont(QFont("Microsoft YaHei", 11))
            is_last = i == len(items) - 1
            color = "#cdd6f4" if is_last else "#6c7086"
            label.setStyleSheet(f"color: {color};")
            if is_last:
                label.setFont(QFont("Microsoft YaHei", 11, QFont.Weight.Bold))
            layout.addWidget(label)
        
        layout.addStretch()


class ToggleSwitch(QFrame):
    """切换开关"""
    def __init__(self, checked=False, parent=None):
        super().__init__(parent)
        self.setFixedSize(48, 26)
        self.checked = checked
        self.update_style()
        
    def update_style(self):
        color = "#3b82f6" if self.checked else "#6c7086"
        self.setStyleSheet(f"""
            ToggleSwitch {{
                background-color: {color};
                border-radius: 13px;
            }}
        """)
        
    def mousePressEvent(self, event):
        self.checked = not self.checked
        self.update_style()


class Tooltip(QLabel):
    """工具提示"""
    def __init__(self, text="", parent=None):
        super().__init__(text, parent)
        self.setFont(QFont("Microsoft YaHei", 10))
        self.setStyleSheet("""
            background-color: #181825;
            color: #cdd6f4;
            border-radius: 6px;
            padding: 6px 10px;
            border: 1px solid #313244;
        """)
        self.setWindowFlags(Qt.WindowType.ToolTip)


class PageHeader(QFrame):
    """页面标题"""
    def __init__(self, title="", subtitle="", parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 16)
        layout.setSpacing(4)
        
        self.title = QLabel(title)
        self.title.setFont(QFont("Microsoft YaHei", 24, QFont.Weight.Bold))
        self.title.setStyleSheet("color: #cdd6f4;")
        layout.addWidget(self.title)
        
        if subtitle:
            self.subtitle = QLabel(subtitle)
            self.subtitle.setFont(QFont("Microsoft YaHei", 12))
            self.subtitle.setStyleSheet("color: #6c7086;")
            layout.addWidget(self.subtitle)


class ActionMenu(QFrame):
    """操作菜单"""
    def __init__(self, actions=None, parent=None):
        super().__init__(parent)
        actions = actions or []
        self.setStyleSheet("""
            ActionMenu {
                background-color: #181825;
                border-radius: 12px;
                border: 1px solid #313244;
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(4)
        
        for icon, text, callback in actions:
            btn = QPushButton(f"{icon}  {text}")
            btn.setFont(QFont("Microsoft YaHei", 11))
            btn.setStyleSheet("""
                QPushButton {
                    background-color: transparent;
                    color: #cdd6f4;
                    border: none;
                    border-radius: 8px;
                    padding: 10px 16px;
                    text-align: left;
                }
                QPushButton:hover {
                    background-color: #313244;
                }
            """)
            if callback:
                btn.clicked.connect(callback)
            layout.addWidget(btn)


class FilterBar(QFrame):
    """筛选栏"""
    def __init__(self, filters=None, parent=None):
        super().__init__(parent)
        filters = filters or []
        self.setStyleSheet("""
            FilterBar {
                background-color: #181825;
                border-radius: 12px;
                border: 1px solid #313244;
            }
        """)
        
        layout = QHBoxLayout(self)
        layout.setContentsMargins(12, 8, 12, 8)
        layout.setSpacing(8)
        
        for label, options in filters:
            combo = QComboBox()
            combo.setFont(QFont("Microsoft YaHei", 11))
            combo.addItem(label)
            combo.addItems(options)
            combo.setStyleSheet("""
                QComboBox {
                    background-color: #1e1e2e;
                    color: #cdd6f4;
                    border: 1px solid #313244;
                    border-radius: 8px;
                    padding: 6px 12px;
                    min-width: 120px;
                }
            """)
            layout.addWidget(combo)
        
        layout.addStretch()


class DataTable(QTableWidget):
    """数据表格"""
    def __init__(self, headers=None, parent=None):
        headers = headers or []
        super().__init__(parent)
        self.setColumnCount(len(headers))
        self.setHorizontalHeaderLabels(headers)
        self.setStyleSheet("""
            DataTable {
                background-color: #181825;
                border: none;
                border-radius: 12px;
                gridline-color: #313244;
            }
            QHeaderView::section {
                background-color: #1e1e2e;
                color: #6c7086;
                padding: 12px;
                border: none;
                font-weight: bold;
            }
            DataTable::item {
                color: #cdd6f4;
                padding: 10px;
                border-bottom: 1px solid #313244;
            }
            DataTable::item:selected {
                background-color: #3b82f630;
            }
        """)
        self.horizontalHeader().setStretchLastSection(True)
        self.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.setAlternatingRowColors(True)


class ChartCard(QFrame):
    """图表卡片"""
    def __init__(self, title="", parent=None):
        super().__init__(parent)
        self.setStyleSheet("""
            ChartCard {
                background-color: #181825;
                border-radius: 16px;
                border: 1px solid #313244;
            }
        """)
        
        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(20, 20, 20, 20)
        
        if title:
            title_label = QLabel(title)
            title_label.setFont(QFont("Microsoft YaHei", 14, QFont.Weight.Bold))
            title_label.setStyleSheet("color: #cdd6f4;")
            self.layout.addWidget(title_label)


class QuickStats(QFrame):
    """快速统计"""
    def __init__(self, stats=None, parent=None):
        super().__init__(parent)
        stats = stats or []
        self.setStyleSheet("""
            QuickStats {
                background-color: #181825;
                border-radius: 12px;
                border: 1px solid #313244;
            }
        """)
        
        layout = QHBoxLayout(self)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(24)
        
        for icon, label, value in stats:
            item = QVBoxLayout()
            item.setSpacing(4)
            
            icon_label = QLabel(icon)
            icon_label.setFont(QFont("Microsoft YaHei", 20))
            item.addWidget(icon_label)
            
            value_label = QLabel(value)
            value_label.setFont(QFont("Microsoft YaHei", 18, QFont.Weight.Bold))
            value_label.setStyleSheet("color: #cdd6f4;")
            item.addWidget(value_label)
            
            label_label = QLabel(label)
            label_label.setFont(QFont("Microsoft YaHei", 10))
            label_label.setStyleSheet("color: #6c7086;")
            item.addWidget(label_label)
            
            layout.addLayout(item)
        
        layout.addStretch()


class NotificationItem(QFrame):
    """通知项"""
    def __init__(self, icon="", title="", time="", unread=True, parent=None):
        super().__init__(parent)
        bg = "#1e1e2e" if unread else "transparent"
        self.setStyleSheet(f"""
            NotificationItem {{
                background-color: {bg};
                border-radius: 10px;
                border-left: 3px solid {'#3b82f6' if unread else 'transparent'};
            }}
        """)
        
        layout = QHBoxLayout(self)
        layout.setContentsMargins(12, 10, 12, 10)
        layout.setSpacing(12)
        
        icon_label = QLabel(icon)
        icon_label.setFont(QFont("Microsoft YaHei", 20))
        layout.addWidget(icon_label)
        
        text = QVBoxLayout()
        text.setSpacing(2)
        
        title_label = QLabel(title)
        title_label.setFont(QFont("Microsoft YaHei", 11))
        title_label.setStyleSheet(f"color: {'#cdd6f4' if unread else '#6c7086'};")
        text.addWidget(title_label)
        
        time_label = QLabel(time)
        time_label.setFont(QFont("Microsoft YaHei", 9))
        time_label.setStyleSheet("color: #6c7086;")
        text.addWidget(time_label)
        
        layout.addLayout(text, 1)


class UserProfile(QFrame):
    """用户资料卡片"""
    def __init__(self, name="", role="", avatar="", parent=None):
        super().__init__(parent)
        self.setStyleSheet("""
            UserProfile {
                background-color: #181825;
                border-radius: 16px;
                border: 1px solid #313244;
            }
        """)
        
        layout = QHBoxLayout(self)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(12)
        
        avatar_label = QLabel(avatar)
        avatar_label.setFont(QFont("Microsoft YaHei", 32))
        layout.addWidget(avatar_label)
        
        text = QVBoxLayout()
        text.setSpacing(2)
        
        name_label = QLabel(name)
        name_label.setFont(QFont("Microsoft YaHei", 14, QFont.Weight.Bold))
        name_label.setStyleSheet("color: #cdd6f4;")
        text.addWidget(name_label)
        
        role_label = QLabel(role)
        role_label.setFont(QFont("Microsoft YaHei", 11))
        role_label.setStyleSheet("color: #6c7086;")
        text.addWidget(role_label)
        
        layout.addLayout(text)
        layout.addStretch()


class SystemMonitor(QFrame):
    """系统监控面板"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet("""
            SystemMonitor {
                background-color: #181825;
                border-radius: 16px;
                border: 1px solid #313244;
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(16)
        
        title = QLabel("⚡ 系统状态")
        title.setFont(QFont("Microsoft YaHei", 14, QFont.Weight.Bold))
        title.setStyleSheet("color: #cdd6f4;")
        layout.addWidget(title)
        
        # CPU
        cpu_layout = QVBoxLayout()
        cpu_header = QHBoxLayout()
        cpu_label = QLabel("CPU")
        cpu_label.setFont(QFont("Microsoft YaHei", 11))
        cpu_label.setStyleSheet("color: #6c7086;")
        cpu_header.addWidget(cpu_label)
        self.cpu_value = QLabel("45%")
        self.cpu_value.setFont(QFont("Microsoft YaHei", 11, QFont.Weight.Bold))
        self.cpu_value.setStyleSheet("color: #3b82f6;")
        cpu_header.addWidget(self.cpu_value)
        cpu_layout.addLayout(cpu_header)
        
        self.cpu_bar = GradientProgressBar()
        self.cpu_bar.setValue(45)
        cpu_layout.addWidget(self.cpu_bar)
        layout.addLayout(cpu_layout)
        
        # Memory
        mem_layout = QVBoxLayout()
        mem_header = QHBoxLayout()
        mem_label = QLabel("内存")
        mem_label.setFont(QFont("Microsoft YaHei", 11))
        mem_label.setStyleSheet("color: #6c7086;")
        mem_header.addWidget(mem_label)
        self.mem_value = QLabel("62%")
        self.mem_value.setFont(QFont("Microsoft YaHei", 11, QFont.Weight.Bold))
        self.mem_value.setStyleSheet("color: #8b5cf6;")
        mem_header.addWidget(self.mem_value)
        mem_layout.addLayout(mem_header)
        
        self.mem_bar = GradientProgressBar()
        self.mem_bar.setValue(62)
        mem_layout.addWidget(self.mem_bar)
        layout.addLayout(mem_layout)


class ActivityFeed(QFrame):
    """活动动态"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet("""
            ActivityFeed {
                background-color: #181825;
                border-radius: 16px;
                border: 1px solid #313244;
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)
        
        title = QLabel("📊 实时动态")
        title.setFont(QFont("Microsoft YaHei", 14, QFont.Weight.Bold))
        title.setStyleSheet("color: #cdd6f4;")
        layout.addWidget(title)
        
        activities = [
            ("🚚", "车辆1完成配送", "2分钟前"),
            ("📦", "新订单 #1234", "5分钟前"),
            ("✅", "订单 #1230 已签收", "12分钟前"),
            ("🔄", "路线优化完成", "15分钟前"),
        ]
        
        for icon, text, time in activities:
            item = QHBoxLayout()
            item.setSpacing(12)
            
            icon_label = QLabel(icon)
            icon_label.setFont(QFont("Microsoft YaHei", 16))
            item.addWidget(icon_label)
            
            text_label = QLabel(text)
            text_label.setFont(QFont("Microsoft YaHei", 11))
            text_label.setStyleSheet("color: #cdd6f4;")
            item.addWidget(text_label, 1)
            
            time_label = QLabel(time)
            time_label.setFont(QFont("Microsoft YaHei", 10))
            time_label.setStyleSheet("color: #6c7086;")
            item.addWidget(time_label)
            
            layout.addLayout(item)
        
        layout.addStretch()


class DashboardWidget(QWidget):
    """仪表板组件"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setup_ui()
        
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)
        
        # 标题
        header = PageHeader("📊 物流调度中心", "实时监控与数据分析")
        layout.addWidget(header)
        
        # 快速统计
        stats = QuickStats([
            ("🚛", "活跃车辆", "12"),
            ("📦", "今日订单", "156"),
            ("✅", "已完成", "89"),
            ("⏱️", "平均时效", "32min"),
        ])
        layout.addWidget(stats)
        
        # 中间区域
        middle = QHBoxLayout()
        middle.setSpacing(20)
        
        # 系统监控
        monitor = SystemMonitor()
        middle.addWidget(monitor)
        
        # 活动动态
        activity = ActivityFeed()
        middle.addWidget(activity)
        
        layout.addLayout(middle)
        
        # 图表区域
        charts = QHBoxLayout()
        charts.setSpacing(20)
        
        chart1 = ChartCard("📈 配送趋势")
        charts.addWidget(chart1, 1)
        
        chart2 = ChartCard("🗺️ 区域分布")
        charts.addWidget(chart2, 1)
        
        layout.addLayout(charts)
        layout.addStretch()


# ====================== 程序入口 ======================
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setFont(QFont("Microsoft YaHei", 10))
    window = LogisticsApp()
    window.show()
    sys.exit(app.exec())